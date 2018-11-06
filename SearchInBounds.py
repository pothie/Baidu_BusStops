import requests
import json
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import *
import tkinter.messagebox
from tkinter.filedialog import askdirectory
from tkinter import ttk

#重置已存在excel的内容时，先人为清空后再运行程序

global tag,PROVS,CITYS,AREAS,pro,ci,ar,ak

tag0 =["美食","酒店","购物","生活服务","丽人","旅游景点","休闲娱乐","运动健身","教育培训","文化传媒","医疗","汽车服务","交通设施","金融","房地产","公司企业","政府机构","出入口","自然地物"]
PROVS = []
pro = ""
ci = ""
ar = ""

ak= "IWS5H16MFHFwQ30p2xDvoXLhrqDypah1"  # 自己的 AK，需要申请
Path = r'' #保存路径，不输入时默认为与本文件统一路径
    
def check(num):
    comma = 0
    if num == None:
        tk.messagebox.showinfo("ERROR","请输入经纬度")
        return 
    for i in range(len(num)-1):
        if num[i] == ',':
            comma = 1
            break
    if comma == 0:
        tk.messagebox.showinfo("ERROR", "请用半角逗号分隔经纬度")
        return

def getline(num):
    for i in range(len(tag00)):
        if num == i:
            line = tag00[i].strip()
            line = line.split('、')
    return line

def location():
    file = open('省市县.json','r')
    global PROVS
    PROVS= ['']
    for item in json.load(file):
        PROVS.append(item['name'])
    file.close()
	
def window():
    window=tk.Tk()
    window.title('地图数据')
    window.geometry('700x300')
	
    frame1 = tk.Frame(window)
    frame1.pack()

    Lak = tk.Label(frame1,text="百度密钥").grid(row=0)
    global ak
    AK = StringVar()
    AK.set(ak)
    Eak = tk.Entry(frame1,width= 40,textvariable = AK)
    Eak.grid(row =0,column=1,columnspan =2)
    def setAk(*args):
        ak = Eak.get()		
    Bak = tk.Button(frame1, text = "确认",command = setAk).grid(row = 0, column = 3)

    Lleft = tk.Label(frame1,text="搜索区左下角经纬度坐标").grid(row=1)
    left = tk.Entry(frame1)
    left.grid(row =1,column=1)
    Rleft = tk.Label(frame1,text="纬度在前，经度在后，用半角逗号隔开").grid(row=1,column=2)
    left.focus()
	
    Lright = tk.Label(frame1,text="搜索区右上角经纬度坐标").grid(row=2)
    right=tk.Entry(frame1)
    right.grid(row=2,column=1)
    Rright = tk.Label(frame1,text="纬度在前，经度在后，用半角逗号隔开").grid(row=2,column=2)
			
    Ltag = tk.Label(frame1,text="搜索关键字").grid(row=3)
    TAGS = ttk.Combobox(frame1,width=17,state='readonly')
    TAGS['values'] = tag0
    TAGS.grid(row=3,column=1) 
    TAGS1 = ttk.Combobox(frame1,width=17,state='readonly')
    TAGS1.grid(row=3,column=2) 
	
    def setTag(*args):		
        TAGS1['values'] = getline(TAGS.current())
    TAGS.bind("<<ComboboxSelected>>",setTag)

    Lloc = tk.Label(frame1,text="搜索地点选择").grid(row=4)
    location()
    provs = ttk.Combobox(frame1,width=17,state='readonly')
    provs.grid(row=4,column=1)
    provs['values'] = PROVS
    
    citys = ttk.Combobox(frame1,width=17,state='readonly')
    citys.grid(row=4,column=2)
	
    areas = ttk.Combobox(frame1,width=17,state='readonly')
    areas.grid(row=4,column=3)
	
    def setCity(*args):
        file = open('省市县.json','r')
        CITYS = ['']
        for item in json.load(file):
            if item['name'] == provs.get():
                for c in item['city']:
                    CITYS.append(c['name'])
        citys['values'] = CITYS
        file.close()
		
    provs.bind("<<ComboboxSelected>>",setCity)
	
    def setArea(*args):
        file = open('省市县.json','r')
        AREAS = ['']
        for item in json.load(file):
            if item['name'] == provs.get():
                for c in item['city']:
                    if c['name'] == citys.get():		
                        AREAS.extend(c['area'])
        file.close()
        areas['values'] = AREAS
    citys.bind("<<ComboboxSelected>>",setArea)
	
    def selectPath():
        path_ = askdirectory()
        path.set(path_)
        global Path
        print(path_ + "," + str(path_))
        Path = path_+"/"

    path = StringVar()		
    Label(frame1,text = "目标路径:").grid(row = 5, column = 0)
    Entry(frame1, textvariable = path).grid(row = 5, column = 1)
    Button(frame1, text = "路径选择", command = selectPath).grid(row = 5, column = 2)
	
    def setVarE():
        global tag
        bLeft = left.get()
        check(bLeft)
        uRight = right.get()
        check(uRight)
		
        if TAGS.get()!=None:
            if TAGS1.get() == None: 
                tag = TAGS.get()
            else:
                tag = TAGS1.get()
        else:
            tk.messagebox.showinfo("ERROR","未选择搜索关键字,请重试")
            return 
        global pro,ci,ar
        pro = provs.get()
        ci = citys.get()
        ar = areas.get()
        request_data_excel(bLeft,uRight)
        tk.messagebox.showinfo("提示","搜索完成")
		
    b1=tk.Button(frame1,text='搜索生成excel',command = setVarE).grid(row=6,column=1)
    #b2=tk.Button(frame1,text='搜索生成txt',command = setVarT).grid(row=3,column=1)
	
    window.mainloop()
				
def request_data_excel(bLeft,uRight):
    global tag,ak,pro,ci,ar
    url = "http://api.map.baidu.com/place/v2/search?query="+tag+"&page_size=20&scope=1&bounds="+bLeft+","+uRight+"&output=json&ak="+ak
    params = {'page_num':0}  
    request = requests.get(url,params=params)  
    time.sleep(0.5)
    try:
        total = json.loads(request.text).get('total') 
    except KeyError:
        tk.messagebox.showinfo("ERROR","百度每日POI访问量限制，正在退出")
        os._exit(0) 
    if total == None or 0:
        tk.messagebox.showinfo("ERROR","搜索地区无此地点")
        return 
		
    if total == 400:
        split(bLeft,uRight)
    elif total > 0:
        total_page_num = (total+19) // 20
        try:
            wb = load_workbook(Path+pro+ci+ar+tag+'.xlsx')
        except FileNotFoundError as error:
            wb = Workbook()
		
        sheet = wb.active
        sheet.title = tag
        sheet['A1'] = 'name'
        sheet['B1'] = 'lng'
        sheet['C1'] = 'lat'
	
        for i in range(total_page_num):
            params['page_num'] = i
            request = requests.get(url,params=params)

            try:
                time.sleep(1)
		        #百度并发量限制，个人或企业认证可提高上限5倍-10倍
		        #根据分割小矩形的大小改变()中的数字，小矩形太多时，百度会提醒并发量过高，请手动提高()中的数字
		        #()中可以是小数
                for item in json.loads(request.text)['results']:
                    if (ar == item['area'] and pro == item['province'] and ci == item['city']) or (ar == "" and ci == item['city'] and pro == item['province']) or (ar == "" and ci == "" and pro == item['province']) or (ar == "" and ci == "" and pro == ""):
                        name = item['name']             
                        lat = item['location']['lat']
                        lng = item['location']['lng']              
                        new_item = (name,lng,lat)
                        sheet.append(new_item)
            except KeyError as error:
                tk.messagebox.showinfo("ERROR","百度每日POI访问量限制，正在退出")
                os._exit(0)
            wb.save(Path+pro+ci+ar+tag+'.xlsx')

def split(bLeft,uRight):
   
    start_lat = "" 
    for i in range(len(bLeft)):
        if bLeft[i] == ',':
            comma = i	
            break
        else:
            start_lat = start_lat + bLeft[i]
			
    start_lng = ""
    for i in range(comma+1,len(bLeft)):
        start_lng = start_lng + bLeft[i]
    
    end_lat = ""
    for i in range(len(uRight)):
        if uRight[i] == ',':
            comma = i	
            break
        else:
            end_lat = end_lat + uRight[i]
			
    end_lng = ""
    for i in range(comma+1,len(uRight)):
        end_lng = end_lng + uRight[i]

	#loop检索
    lat = (float(start_lat) + float(end_lat))/2
    lng = (float(start_lng) + float(end_lng))/2

    print(start_lat+","+start_lng,str(lat)+","+str(lng))
    request_data_excel(start_lat+","+start_lng,str(lat)+","+str(lng))
    print(start_lat+","+str(lng),str(lat)+","+end_lng)
    request_data_excel(start_lat+","+str(lng),str(lat)+","+end_lng)
    print(str(lat)+","+start_lng,end_lat+","+str(lng))
    request_data_excel(str(lat)+","+start_lng,end_lat+","+str(lng))
    print(str(lat)+","+str(lng),end_lat+","+end_lng)
    request_data_excel(str(lat)+","+str(lng),end_lat+","+end_lng)
    
if __name__ == '__main__':
    
    f = open('关键字.txt','r')
    tag00 = f.readlines()
    f.close()
    window()